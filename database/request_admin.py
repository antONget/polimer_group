import sqlite3
from openpyxl import load_workbook
import os

def create_tables():
    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute(
        f'CREATE TABLE IF NOT EXISTS ground_tanks (name_razdel text,name text,art text,size text,weight text,volume int,cost_mitichi int,cost_zuevo int,image_name text)')
    conn.commit()
    cur.execute(
        f'CREATE TABLE IF NOT EXISTS underground_tanks (name_razdel text,name text,art text,size text,weight text,volume int,cost_mitichi int,cost_zuevo int,image_name text)')
    conn.commit()
    cur.execute(
        f'CREATE TABLE IF NOT EXISTS for_village (name_razdel text,name text,art text,size text,weight text,volume int,cost_mitichi int,cost_zuevo int,image_name text)')
    conn.commit()
    cur.execute(
        'CREATE TABLE IF NOT EXISTS accessories (name_razdel text,name text,art text,cost_mitichi text,cost_zuevo text,image_name text)')
    conn.commit()
    cur.execute(
        'CREATE TABLE IF NOT EXISTS for_trash (name_razdel text,name text,art text,harakteristik text,weight text,cost_mitichi int,image_name text)')
    conn.commit()
    cur.execute(
        'CREATE TABLE IF NOT EXISTS boxes (name_razdel text,name text,art text,volume int,size text,weight text,cost_mitichi int,cost_zuevo int,image_name text)')
    conn.commit()
    cur.execute(
        'CREATE TABLE IF NOT EXISTS azs (name_razdel text,name text,volume int,size text,art_piusi text,cost_piusi text,art_belak text,cost_belak text,art_china_premium text,cost_china_premium text,art_china text,cost_china text)')
    conn.commit()
    cur.execute(f'CREATE TABLE IF NOT EXISTS azs_parts (name_razdel text,name text,art text,cost int)')
    conn.commit()
    cur.close()
    conn.close()


def process_excel_file(file_path):
    workbook = load_workbook(file_path)

    sheet = workbook['- Емкости наземные -']

    name_razdel = None

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM ground_tanks')

    for row in range(7,sheet.max_row+1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            art = sheet.cell(row=row, column=2).value
            size = sheet.cell(row=row, column=3).value
            weight = sheet.cell(row=row, column=4).value
            obiem = sheet.cell(row=row, column=5).value
            Mitishi_cost = sheet.cell(row=row, column=6).value
            Orehovo_cost = sheet.cell(row=row, column=7).value
            image_name = sheet.cell(row=row,column=8).value
            cur.execute(f'INSERT INTO ground_tanks (name_razdel,name,art,size,weight,volume,cost_mitichi,cost_zuevo,image_name)'
                        f'VALUES (?,?,?,?,?,?,?,?,?)',(name_razdel,name,art,size,weight,obiem,Mitishi_cost,Orehovo_cost,image_name))
            conn.commit()
    cur.close()
    conn.close()

    sheet = workbook['- Емкости подземные -']

    name_razdel = None

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM underground_tanks')

    for row in range(7, sheet.max_row+1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            art = sheet.cell(row=row, column=2).value
            size = sheet.cell(row=row, column=3).value
            weight = sheet.cell(row=row, column=4).value
            obiem = sheet.cell(row=row, column=5).value
            Mitishi_cost = sheet.cell(row=row, column=6).value
            Orehovo_cost = sheet.cell(row=row, column=7).value
            image_name = sheet.cell(row=row, column=8).value

            cur.execute(f'INSERT INTO underground_tanks (name_razdel,name,art,size,weight,volume,cost_mitichi,cost_zuevo,image_name)'
                        f'VALUES (?,?,?,?,?,?,?,?,?)',(name_razdel,name,art,size,weight,obiem,Mitishi_cost,Orehovo_cost,image_name))
            conn.commit()
    cur.close()
    conn.close()

    sheet = workbook['- Для дачи -']

    name_razdel = None

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM for_village')

    for row in range(7, sheet.max_row+1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            art = sheet.cell(row=row, column=2).value
            size = sheet.cell(row=row, column=3).value
            weight = sheet.cell(row=row, column=4).value
            Mitishi_cost = sheet.cell(row=row, column=5).value
            Orehovo_cost = sheet.cell(row=row, column=6).value
            image_name = sheet.cell(row=row,column=7).value

            cur.execute(f'INSERT INTO for_village (name_razdel,name,art,size,weight,cost_mitichi,cost_zuevo,image_name)'
                        f'VALUES (?,?,?,?,?,?,?,?)',(name_razdel,name,art,size,weight,Mitishi_cost,Orehovo_cost,image_name))
            conn.commit()
    cur.close()
    conn.close()

    sheet = workbook['- Комплектующие -']

    name_razdel = 'Комлектующие'

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM accessories')

    for row in range(7, sheet.max_row+1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            art = sheet.cell(row=row, column=2).value
            Mitishi_cost = sheet.cell(row=row, column=3).value
            Orehovo_cost = sheet.cell(row=row, column=4).value
            image_name = sheet.cell(row=row,column=5).value

            cur.execute(f'INSERT INTO accessories (name_razdel,name,art,cost_mitichi,cost_zuevo,image_name) '
                        f' VALUES (?,?,?,?,?,?)',(name_razdel,name,art,Mitishi_cost,Orehovo_cost,image_name))
            conn.commit()
    cur.close()
    conn.close()

    sheet = workbook['- Мусоросбросы -']

    name_razdel = None

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM for_trash')

    for row in range(7, sheet.max_row + 1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            art = sheet.cell(row=row, column=2).value
            haracteristics = sheet.cell(row=row, column=3).value
            weight = sheet.cell(row=row, column=4).value
            Mitishi_cost = sheet.cell(row=row, column=5).value
            image_name = sheet.cell(row=row,column=6).value

            cur.execute(f'INSERT INTO for_trash (name_razdel,name,art,harakteristik,weight,cost_mitichi,image_name)'
                        f'VALUES (?,?,?,?,?,?,?)',(name_razdel,name,art,haracteristics,weight,Mitishi_cost,image_name))
            conn.commit()
    cur.close()
    conn.close()

    sheet = workbook['- Ящики -']

    name_razdel = None

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM boxes')

    for row in range(7, sheet.max_row + 1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            art = sheet.cell(row=row, column=2).value
            volume = sheet.cell(row=row, column=3).value
            size = sheet.cell(row=row, column=4).value
            weight = sheet.cell(row=row, column=5).value
            Mitishi_cost = sheet.cell(row=row, column=6).value
            Orehovo_cost = sheet.cell(row=row, column=7).value
            image_name = sheet.cell(row=row,column=8).value

            cur.execute(f'INSERT INTO boxes (name_razdel,name,art,volume,size,weight,cost_mitichi,cost_zuevo,image_name)'
                        f'VALUES (?,?,?,?,?,?,?,?,?)',(name_razdel,name,art,volume,size,weight,Mitishi_cost,Orehovo_cost,image_name))
            conn.commit()
    cur.close()
    conn.close()

    sheet = workbook['- Мини АЗС -']

    name_razdel = None

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM azs')

    for row in range(7,sheet.max_row+1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            volume = sheet.cell(row=row, column=2).value
            size = sheet.cell(row=row, column=3).value
            art_piusi = sheet.cell(row=row, column=4).value
            cost_piusi = sheet.cell(row=row, column=5).value
            art_belak = sheet.cell(row=row, column=6).value
            cost_belak = sheet.cell(row=row, column=7).value
            art_china_premium = sheet.cell(row=row, column=8).value
            cost_china_remium = sheet.cell(row=row, column=9).value
            art_china = sheet.cell(row=row, column=10).value
            cost_china = sheet.cell(row=row, column=11).value
            cur.execute(
                f'INSERT INTO azs (name_razdel,name,volume,size,art_piusi,cost_piusi,art_belak,cost_belak,art_china_premium,cost_china_premium,art_china,cost_china)'
                f'VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                (name_razdel,name, volume, size, art_piusi, cost_piusi, art_belak, cost_belak, art_china_premium, cost_china_remium,
                 art_china, cost_china))
            conn.commit()
    cur.close()
    conn.close()

    sheet = workbook['- Комплектующие для АЗС -']

    name_razdel = None

    conn = sqlite3.connect('database/TOVARI.sql')
    cur = conn.cursor()
    cur.execute('DELETE FROM azs_parts')



    for row in range(7,sheet.max_row+1):
        if ((sheet.cell(row=row, column=2).value == None) and (sheet.cell(row=row, column=3).value == None) and
            (sheet.cell(row=row, column=4).value == None) and(sheet.cell(row=row, column=5).value == None)):
            name_razdel = sheet.cell(row=row, column=1).value
        else:
            name = sheet.cell(row=row, column=1).value
            art = sheet.cell(row=row, column=2).value
            cost = sheet.cell(row=row, column=3).value
            cur.execute(f'INSERT INTO azs_parts (name_razdel,name,art,cost) VALUES (?,?,?,?)', (name_razdel,name,art,cost))
            conn.commit()
    cur.close()
    conn.close()


def get_photo_to_sqlite(photo_name,photo_id):
    conn = sqlite3.connect('database/IMAGES_IDS.sql')
    cur = conn.cursor()
    cur.execute('INSERT INTO image_id (name,id) VALUES (?,?)',(photo_name,photo_id))
    conn.commit()
    cur.close()
    conn.close()

def get_all_photo_name():
    conn = sqlite3.connect('database/IMAGES_IDS.sql')
    cur = conn.cursor()
    cur.execute('SELECT name FROM image_id')
    data = cur.fetchall()
    cur.close()
    conn.close()
    return data

def rechange_photo_id(photo_name,photo_id):
    conn = sqlite3.connect('database/IMAGES_IDS.sql')
    cur = conn.cursor()
    cur.execute(f'UPDATE image_id SET id = ? WHERE name = ?',(photo_name,photo_id))
    conn.commit()
    cur.close()
    conn.close()